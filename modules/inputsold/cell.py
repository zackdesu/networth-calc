'cell name'

from openpyxl.utils import get_column_letter


def cell(ws, rownum=11, colnum=4):  # pylint: disable=invalid-name
    'cell'
    empty_cell = []
    for row in range(1, rownum):
        for col in range(1, colnum):
            letnum = get_column_letter(col) + str(row)
            cells = ws[letnum].value
            if cells is None:
                empty_cell.append(letnum)
    return empty_cell
